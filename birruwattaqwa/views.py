from datetime import date, datetime, time, timedelta
import math
import qrcode
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import (login_required, permission_required,user_passes_test)
from django.contrib.auth.models import Group, User
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .forms import (AdminCreateUserForm,AbsensiForm,AbsensiManualForm,ProfilGuruForm)
from .models import (Absensi,ProfilGuru,LokasiAbsen)
from django.http import HttpResponse
import json
from django.http import JsonResponse
from django.utils.dateparse import parse_time


def haversine(lat1, lon1, lat2, lon2):
    R = 6371  # km
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c * 1000  # meter


from .models import LokasiAbsen  # pastikan ini di-import

@login_required
def absen_guru(request):
    user = request.user
    now_time = now().time()
    today = now().date()
    bulan, tahun = now().strftime('%B'), now().year
    qr_code_value = f"ABSEN-{today}"

    form = AbsensiForm()
    batas_absen = time(9, 0)

    # Otomatis dianggap Alfa jika belum absen dan sudah lewat batas waktu
    if now_time > batas_absen:
        sudah_absen = Absensi.objects.filter(guru=user, tanggal=today).exists()
        if not sudah_absen:
            Absensi.objects.create(
                guru=user,
                tanggal=today,
                status="Alfa",
                keterangan="Tidak Hadir (Alpha)",
                bulan=bulan,
                tahun=tahun,
                jam_absensi=now().time(),
            )
            messages.warning(request, "Waktu absensi sudah lewat. Anda dianggap Alfa hari ini.")
            return redirect('absent')

    if request.method == 'POST':
        form = AbsensiForm(request.POST)

        if form.is_valid():
            status = form.cleaned_data['status']
            keterangan = request.POST.get('keterangan', '').strip() if status in ["Sakit", "Izin"] else "Tidak Hadir (Alpha)"

            # Jika status Sakit atau Izin, langsung catat absensi tanpa cek lokasi
            if status in ["Sakit", "Izin"]:
                Absensi.objects.create(
                    guru=user,
                    tanggal=today,
                    status=status,
                    keterangan=keterangan,
                    bulan=bulan,
                    tahun=tahun,
                    jam_absensi=now().time(),
                )
                messages.success(request, f"Absensi {status} berhasil diajukan.")
                return render(request, 'birruwattaqwa/guru/absent.html', {
                    'qr_code_value': qr_code_value,
                    'form': form,
                })

            # Jika status Hadir, lakukan pengecekan lokasi
            elif status == "Hadir":
                latitude = request.POST.get('latitude', '').strip()
                longitude = request.POST.get('longitude', '').strip()

                if latitude and longitude:
                    try:
                        latitude = float(latitude)
                        longitude = float(longitude)

                        lokasi = LokasiAbsen.objects.first()  # ambil lokasi dari DB
                        if not lokasi:
                            messages.error(request, "Lokasi absensi belum dikonfigurasi oleh admin.")
                        else:
                            distance = haversine(latitude, longitude, lokasi.latitude, lokasi.longitude)

                            if distance > lokasi.radius_meter:
                                messages.error(request, f"Anda berada di luar area absensi! Jarak: {distance:.2f} meter")
                            else:
                                Absensi.objects.create(
                                    guru=user,
                                    tanggal=today,
                                    status=status,
                                    keterangan="Hadir",
                                    bulan=bulan,
                                    tahun=tahun,
                                    jam_absensi=now().time(),
                                    latitude=latitude,
                                    longitude=longitude
                                )
                                messages.success(request, "Absensi berhasil.")
                                return redirect('dashboard_guru')

                    except ValueError:
                        messages.error(request, "Lokasi tidak valid. Pastikan GPS aktif dan coba lagi.")
                else:
                    messages.error(request, "Lokasi tidak terdeteksi. Pastikan GPS aktif.")

    return render(request, 'birruwattaqwa/guru/absent.html', {
        'qr_code_value': qr_code_value,
        'form': form,
    })

    
def tandai_guru_yang_alfa_pagi():
    today = now().date()
    bulan = now().strftime('%B')
    tahun = now().year
    current_time = now().time()
    batas_waktu = time(9, 0)  # batas jam 9 pagi

    if current_time >= batas_waktu:
        semua_guru = User.objects.filter(groups__name='Guru')
        absensi_hari_ini = Absensi.objects.filter(tanggal=today)
        guru_sudah_absen = absensi_hari_ini.values_list('guru__id', flat=True)
        guru_yang_belum_absen = semua_guru.exclude(id__in=guru_sudah_absen)

        for guru in guru_yang_belum_absen:
            sudah_alfa = Absensi.objects.filter(guru=guru, tanggal=today, status="Alfa").exists()
            if not sudah_alfa:
                Absensi.objects.create(
                    guru=guru,
                    tanggal=today,
                    status="Alfa",
                    keterangan="Tidak Hadir (Alpha)",
                    bulan=bulan,
                    tahun=tahun,
                    jam_absensi=now().time(),
                )

@login_required
def scan_qr(request, qr_code):
    """Verifikasi QR Code untuk mencatat kehadiran"""
    today = now().strftime('%Y-%m-%d')
    expected_code = f"ABSEN-{today}"

    if qr_code == expected_code:
        Absensi.objects.create(
            guru=request.user,
            tanggal=now().date(),
            status="Hadir",
            bulan=now().strftime('%B'),
            tahun=now().year,
            jam_absensi=now().time(),
            lokasi="Sekolah"
        )
        return JsonResponse({"message": "Absensi berhasil!"})
    
    return JsonResponse({"error": "QR Code tidak valid!"})

from django.contrib.auth.models import User  # Import model User

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def view_absensi(request):
    user = request.user
    guru_filter = request.GET.get('guru', '')
    search_query = request.GET.get('search', '')
    bulan_filter = request.GET.get('bulan')
    tahun_filter = request.GET.get('tahun')
    bulan_list = [
    ("01", "Januari"),
    ("02", "Februari"),
    ("03", "Maret"),
    ("04", "April"),
    ("05", "May"),
    ("06", "Juni"),
    ("07", "Juli"),
    ("08", "Agustus"),
    ("09", "September"),
    ("10", "Oktober"),
    ("11", "November"),
    ("12", "Desember"),
]
    tahun_list = [str(tahun) for tahun in range(2023, datetime.date.today().year + 1)]

    if user.groups.filter(name='Admin').exists():
        absensi_list = Absensi.objects.all()
        guru_list = User.objects.filter(groups__name='Guru')

        if guru_filter:
            absensi_list = absensi_list.filter(guru__id=guru_filter)
        if search_query:
            absensi_list = absensi_list.filter(keterangan__icontains=search_query)
        if bulan_filter and tahun_filter:
            absensi_list = absensi_list.filter(tanggal__month=bulan_filter, tanggal__year=tahun_filter)
    else:
        absensi_list = Absensi.objects.filter(guru=user)
        guru_list = None
        if search_query:
            absensi_list = absensi_list.filter(keterangan__icontains=search_query)
        if bulan_filter and tahun_filter:
            absensi_list = absensi_list.filter(tanggal__month=bulan_filter, tanggal__year=tahun_filter)

    # Export ke Excel
    if 'export' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=rekap_absensi_{datetime.date.today()}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Absensi"

        headers = ['Nama', 'Tanggal', 'Status', 'Keterangan']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        ws.append(headers)
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for absensi in absensi_list:
            ws.append([
                absensi.guru.username,
                absensi.tanggal.strftime("%Y-%m-%d"),
                absensi.status,
                absensi.keterangan
            ])

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        wb.save(response)
        return response

    return render(request, 'birruwattaqwa/admin/list_absen.html', {
        'absensi_list': absensi_list,
        'guru_list': guru_list,
        'selected_guru': guru_filter,
        'search_query': search_query,
        'selected_bulan': bulan_filter,
        'selected_tahun': tahun_filter,
        'bulan_list': bulan_list,
        'tahun_list': tahun_list,
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def edit_absensi(request, absensi_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            absensi = Absensi.objects.get(id=absensi_id)
            absensi.status = data.get('status')
            absensi.keterangan = data.get('keterangan')
            absensi.save()
            return JsonResponse({'success': True, 'message': 'Absensi berhasil diupdate.'})
        except Absensi.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Absensi tidak ditemukan.'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data.'}, status=400)
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan.'}, status=405)



@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def delete_absensi(request, absensi_id):
    if request.method == 'POST':
        try:
            absensi = Absensi.objects.get(id=absensi_id)
            absensi.delete()
            return JsonResponse({'success': True, 'message': 'Absensi berhasil dihapus.'})
        except Absensi.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Absensi tidak ditemukan.'}, status=404)
    return JsonResponse({'success': False, 'message': 'Metode tidak diizinkan.'}, status=405)


from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from .models import Absensi, User  # Pastikan model sudah sesuai

@login_required
def dashboard_admin(request):
    # Dapatkan hari ini
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    # Filter absensi 7 hari terakhir
    absensi_minggu_ini = Absensi.objects.filter(tanggal__range=[seven_days_ago, today])

    # Buat dictionary: tanggal -> jumlah status
    kehadiran_per_hari = {}
    for i in range(7):
        hari = seven_days_ago + timedelta(days=i)
        kehadiran_per_hari[hari.strftime('%d-%m')] = {'Hadir': 0, 'Izin': 0, 'Sakit': 0, 'Alfa': 0}

    for a in absensi_minggu_ini:
        key = a.tanggal.strftime('%d-%m')
        if a.status in kehadiran_per_hari[key]:
            kehadiran_per_hari[key][a.status] += 1

    # Ambil data untuk grafik
    labels = list(kehadiran_per_hari.keys())
    hadir_data = [kehadiran_per_hari[tgl]['Hadir'] for tgl in labels]
    izin_data = [kehadiran_per_hari[tgl]['Izin'] for tgl in labels]
    sakit_data = [kehadiran_per_hari[tgl]['Sakit'] for tgl in labels]
    alfa_data = [kehadiran_per_hari[tgl]['Alfa'] for tgl in labels]

    # Ambil total jumlah guru dan absensi hari ini
    total_guru = User.objects.filter(groups__name='Guru').count()
    total_hadir = Absensi.objects.filter(tanggal=today, status='Hadir').count()
    total_izin = Absensi.objects.filter(tanggal=today, status='Izin').count()
    total_sakit = Absensi.objects.filter(tanggal=today, status='Sakit').count()
    total_alfa = Absensi.objects.filter(tanggal=today, status='Alfa').count()

    # Konteks yang akan diteruskan ke template
    context = {
        # Data untuk grafik
        'labels': labels,
        'hadir_data': hadir_data,
        'izin_data': izin_data,
        'sakit_data': sakit_data,
        'alfa_data': alfa_data,
        # Data untuk statistik kecil
        'total_guru': total_guru,
        'total_izin': total_izin,
        'total_sakit': total_sakit,
        'total_alfa': total_alfa,
        'total_hadir' : total_hadir
    }

    return render(request, 'birruwattaqwa/admin/dashboard_admin.html', context)

from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta


@login_required
def dashboard_guru(request):
    today = timezone.now().date()
    bulan_ini = today.replace(day=1)
    tujuh_hari_lalu = today - timedelta(days=6)

    # Ambil absensi user saat ini
    absensi_user = Absensi.objects.filter(guru=request.user)

    # Statistik bulan ini
    total_hadir_bulan_ini = absensi_user.filter(status='Hadir', tanggal__gte=bulan_ini).count()
    total_izin_bulan_ini = absensi_user.filter(status='Izin', tanggal__gte=bulan_ini).count()
    total_sakit_bulan_ini = absensi_user.filter(status='Sakit', tanggal__gte=bulan_ini).count()
    total_alpha_bulan_ini = absensi_user.filter(status='Alfa', tanggal__gte=bulan_ini).count()

    # Hari aktif total
    total_hari_aktif = absensi_user.count()

    context = {
        'total_hadir_bulan_ini': total_hadir_bulan_ini,
        'total_izin_bulan_ini': total_izin_bulan_ini,
        'total_sakit_bulan_ini': total_sakit_bulan_ini,
        'total_alpha_bulan_ini': total_alpha_bulan_ini,
        'total_hari_aktif': total_hari_aktif,
    }

    return render(request, 'birruwattaqwa/guru/dashboard_guru.html', context)


@login_required
def redirect_dashboard(request):
    if request.user.groups.filter(name='Admin').exists():
        return redirect('dashboard_admin')
    elif request.user.groups.filter(name='Guru').exists():
        return redirect('dashboard_guru')
    else:
        return redirect('home')  # Jika role tidak terdaftar

@csrf_exempt
def login_guru(request):
    """Halaman Login: Setelah login, diarahkan ke halaman absen"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Username atau password salah.')

    return render(request, 'registrasi/login.html')



def logout_guru(request):
    """Logout dan kembali ke halaman home"""
    logout(request)
    return redirect('home')


def home(request):
    """Halaman Home: Bisa diakses semua orang"""
    
    return render(request, 'home.html')



@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())  # Hanya Admin bisa akses
def create_user(request):
    if request.method == 'POST':
        form = AdminCreateUserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            print("User akan disimpan:", user.username, user.email)
            user.save()

            # Tambahkan ke grup (buat jika belum ada)
            group, created = Group.objects.get_or_create(name=form.cleaned_data['role'])
            print("Grup ditemukan atau dibuat:", group)
            user.groups.add(group)

            # Simpan data alamat dan jabatan ke model ProfilGuru
            ProfilGuru.objects.create(
                user=user,

                jabatan=form.cleaned_data['jabatan']
            )
            print("ProfilGuru berhasil dibuat")

            return redirect('list_users')  # Redirect ke daftar pengguna
        else:
            print("Form tidak valid:", form.errors)
    else:
        form = AdminCreateUserForm()
    
    return render(request, 'birruwattaqwa/admin/registrasi.html', {'form': form})



@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def list_users(request):
    users = User.objects.all()
    jabatan_list = (
        User.objects.exclude(profilguru__jabatan__isnull=True)
        .exclude(profilguru__jabatan__exact='')
        .values_list('profilguru__jabatan', flat=True)
        .distinct()
    )
    return render(request, 'birruwattaqwa/admin/list_users.html', {'users': users, 'jabatan_list': jabatan_list})


@csrf_exempt
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def edit_user(request, user_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get('username')
            email = data.get('email')
            jabatan = data.get('jabatan')
            role = data.get('role')
            password = data.get('password')  # Tambahkan ini

            user = User.objects.get(id=user_id)

            if username:
                user.username = username
            if email:
                user.email = email
            if password:
                user.set_password(password)  # Ganti password
            user.save()

            # Update ProfilGuru
            try:
                profil = user.profilguru
            except ProfilGuru.DoesNotExist:
                profil = ProfilGuru(user=user)
            if jabatan:
                profil.jabatan = jabatan
            profil.save()

            # Update role (Group)
            if role:
                user.groups.clear()
                group, created = Group.objects.get_or_create(name=role)
                user.groups.add(group)

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@csrf_exempt
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def delete_user(request, user_id):
    if request.method == "POST":
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            return JsonResponse({'success': True})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})
# views.py

from io import BytesIO
from django.core.files.base import ContentFile
from django.shortcuts import render
from django.utils import timezone
from .models import DailyQRCode
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.utils.timezone import localdate
from django.contrib import messages
from django.shortcuts import redirect
@login_required
def generate_daily_qrcode(request, tanggal):
    if request.user.groups.filter(name='Guru').exists():
        try:
            tanggal_qr = datetime.datetime.strptime(tanggal, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "❌ Format tanggal tidak valid.")
            return redirect('scan_qr_view')

        tanggal_hari_ini = localdate()
        if tanggal_qr != tanggal_hari_ini:
            messages.error(request, "❌ QR code ini bukan untuk hari ini.")
            return redirect('scan_qr_view')

        try:
            lat = float(request.GET.get("lat"))
            lon = float(request.GET.get("lon"))
        except (TypeError, ValueError):
            messages.error(request, "❌ Lokasi tidak valid atau tidak tersedia.")
            return redirect('scan_qr_view')

        lokasi = LokasiAbsen.objects.first()
        if not lokasi:
            messages.error(request, "❌ Lokasi absensi belum diatur oleh admin.")
            return redirect('scan_qr_view')

        jarak = haversine(lat, lon, lokasi.latitude, lokasi.longitude)
        if jarak > lokasi.radius_meter:
            messages.warning(request, f"📍 Kamu terlalu jauh dari lokasi absen. (Jarakmu: {int(jarak)} m)")
            return redirect('scan_qr_view')

        if not Absensi.objects.filter(guru=request.user, tanggal=tanggal_hari_ini).exists():
            Absensi.objects.create(
                guru=request.user,
                tanggal=tanggal_hari_ini,
                tahun=tanggal_hari_ini.year,
                bulan=tanggal_hari_ini.strftime("%B"),
                status="Hadir",
                keterangan="Absen via QR"
            )
            messages.success(request, "✅ Absensi berhasil tercatat.")
        else:
            messages.info(request, "⚠️ Kamu sudah absen hari ini.")

        return redirect('scan_qr_view')

    messages.error(request, "❌ Hanya guru yang bisa absen lewat QR code.")
    return redirect('scan_qr_view')


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())  # Cuma admin bisa akses
def generate_admin_qrcode(request):
    today = timezone.localdate()
    daily_qr, created = DailyQRCode.objects.get_or_create(tanggal=today)
    if created or not daily_qr.qrcode:
        # Gunakan URL lokal atau URL absensi yang valid
        url = request.build_absolute_uri(f"/absensi/{today}/")
        qr_img = qrcode.make(url)
        buffer = BytesIO()
        qr_img.save(buffer, format='PNG')
        daily_qr.qrcode.save(f'qr_{today}.png', ContentFile(buffer.getvalue()))
        daily_qr.save()
    return render(request, 'birruwattaqwa/admin/admin_qrcode.html', {'daily_qr': daily_qr})

@login_required
def scan_qr_view(request):
    return render(request, 'birruwattaqwa/guru/qrcodes.html')

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
import datetime
from .models import Absensi

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

from .models import Absensi

@login_required
def rekap_absensi_guru(request):
    user = request.user

    if not user.groups.filter(name='Guru').exists():
        return HttpResponse("Unauthorized", status=403)

    search_query = request.GET.get('search', '')
    bulan_filter = request.GET.get('bulan')
    tahun_filter = request.GET.get('tahun')

    bulan_list = [
        ("01", "Januari"), ("02", "Februari"), ("03", "Maret"),
        ("04", "April"), ("05", "Mei"), ("06", "Juni"),
        ("07", "Juli"), ("08", "Agustus"), ("09", "September"),
        ("10", "Oktober"), ("11", "November"), ("12", "Desember"),
    ]
    tahun_list = [str(tahun) for tahun in range(2023, datetime.date.today().year + 1)]

    absensi_list = Absensi.objects.filter(guru=user)

    if search_query:
        absensi_list = absensi_list.filter(keterangan__icontains=search_query)
    if bulan_filter and tahun_filter:
        absensi_list = absensi_list.filter(
            tanggal__month=int(bulan_filter),
            tanggal__year=int(tahun_filter)
        )

    # Export ke Excel
    if 'export' in request.GET:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=rekap_absensi_{user.username}_{datetime.date.today()}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Absensi"

        headers = ['Tanggal', 'Status', 'Keterangan']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        ws.append(headers)
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for absensi in absensi_list:
            ws.append([
                absensi.tanggal.strftime("%Y-%m-%d"),
                absensi.status,
                absensi.keterangan
            ])

        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        wb.save(response)
        return response

    return render(request, 'birruwattaqwa/guru/rekap_guru.html', {
        'absensi_list': absensi_list,
        'search_query': search_query,
        'selected_bulan': bulan_filter,
        'selected_tahun': tahun_filter,
        'bulan_list': bulan_list,
        'tahun_list': tahun_list,
    })




def simple_view(request):
    return HttpResponse("Hello Ngrok!")

from .forms import LokasiAbsenForm

@login_required
def atur_lokasi(request):
    lokasi, created = LokasiAbsen.objects.get_or_create(
        id=1,
        defaults={
            'nama_tempat': 'Sekolah',
            'latitude': -6.923535,
            'longitude':110.568171,
            'radius_meter': 10,
        }
    )

    if request.method == 'POST':
        form = LokasiAbsenForm(request.POST, instance=lokasi)
        if form.is_valid():
            form.save()
            messages.success(request, "Lokasi berhasil diperbarui.")
            return redirect('atur_lokasi_absen')
    else:
        form = LokasiAbsenForm(instance=lokasi)

    return render(request, 'birruwattaqwa/admin/atur_lokasi.html', {
        'form': form
    })
    




from django.utils.timezone import now
@login_required
def input_absensi_manual(request):
    if request.method == 'POST':
        form = AbsensiManualForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('input_absensi_manual')
        # jika form tidak valid, jangan redirect, render ulang dengan form berisi error
    else:
        form = AbsensiManualForm()

    context = {
        'form': form,
        'guru_list': User.objects.filter(groups__name='Guru'),
        'today': now().date(),
        'now': now()
    }
    return render(request, 'birruwattaqwa/admin/input_manual.html', context)










